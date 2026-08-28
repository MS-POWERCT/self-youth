#!/usr/bin/env python3
"""Generate recognizable farm SVG icons + Excel manifest in user order."""

import re
import sys
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent / ".pylibs"))
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "public" / "images" / "farm" / "icons"
EXCEL = ROOT / "public" / "images" / "farm" / "farm-icons.xlsx"
TWEMOJI = "https://cdn.jsdelivr.net/gh/twitter/twemoji@14.0.2/assets/svg/{}.svg"

# 按用户给定顺序：名称, 文件名, twemoji代码(无则走自定义)
ORDER = [
    ("小麦", "wheat", "1f33e"),
    ("胡萝卜", "carrot", "1f955"),
    ("土豆", "potato", "1f954"),
    ("玉米", "corn", "1f33d"),
    ("番茄", "tomato", "1f345"),
    ("辣椒", "chili", "1f336"),
    ("黄瓜", "cucumber", "1f952"),
    ("大豆", "soybean", "1fad8"),
    ("花生", "peanut", "1f95c"),
    ("南瓜", "pumpkin", None),
    ("洋葱", "onion", None),
    ("茄子", "eggplant", "1f346"),
    ("西瓜", "watermelon", "1f349"),
    ("甘蔗", "sugarcane", None),
    ("薰衣草", "lavender", None),
    ("面粉", "flour", None),
    ("面包", "bread", "1f35e"),
    ("玉米面", "cornmeal", None),
    ("焦糖爆米花", "caramel-popcorn", "1f37f"),
    ("土豆片", "potato-chips", None),
    ("薯条", "french-fries", "1f35f"),
    ("蕃茄酱", "ketchup", None),
    ("番茄意面", "tomato-pasta", "1f35d"),
    ("辣椒粉", "chili-powder", None),
    ("辣酱", "hot-sauce", None),
    ("腌黄瓜", "pickled-cucumber", None),
    ("酸黄瓜沙拉", "pickle-salad", "1f957"),
    ("胡萝卜汁", "carrot-juice", None),
    ("营养麦片", "cereal", None),
    ("豆腐", "tofu", None),
    ("豆腐干", "dried-tofu", None),
    ("花生酱", "peanut-butter", None),
    ("花生饼干", "peanut-cookie", "1f36a"),
    ("南瓜泥", "pumpkin-puree", None),
    ("南瓜派", "pumpkin-pie", "1f967"),
    ("洋葱汤料", "onion-soup-mix", None),
    ("洋葱汤", "onion-soup", "1f372"),
    ("茄子片", "eggplant-slice", None),
    ("茄子煲", "eggplant-casserole", None),
    ("西瓜汁", "watermelon-juice", None),
    ("西瓜冰沙", "watermelon-smoothie", None),
    ("糖浆", "syrup", "1f36f"),
    ("糖果", "candy", "1f36c"),
    ("精油", "essential-oil", None),
    ("香水", "perfume", None),
    ("草莓", "strawberry", "1f353"),
    ("蓝莓", "blueberry", None),
    ("苹果", "apple", "1f34e"),
    ("芒果", "mango", "1f96d"),
    ("葡萄", "grape", "1f347"),
    ("草莓酱", "strawberry-jam", None),
    ("草莓蛋糕", "strawberry-cake", "1f370"),
    ("蓝莓酱", "blueberry-jam", None),
    ("蓝莓马芬", "blueberry-muffin", "1f9c1"),
    ("苹果片", "apple-slice", None),
    ("苹果派", "apple-pie", None),
    ("芒果汁", "mango-juice", None),
    ("芒果冰沙", "mango-smoothie", None),
    ("葡萄汁", "grape-juice", None),
    ("葡萄酒", "wine", "1f377"),
    ("鸡蛋", "egg", "1f95a"),
    ("牛奶", "milk", "1f95b"),
    ("糖", "sugar", None),
    ("油", "oil", None),
    ("花瓣", "petal", "1f338"),
]

SVG_BG = ""

CUSTOM = {
    "pumpkin": '''
<ellipse cx="32" cy="38" rx="20" ry="17" fill="#E87511"/>
<ellipse cx="32" cy="38" rx="20" ry="17" fill="none" stroke="#C65D00" stroke-width="1.2"/>
<path d="M18 36 Q32 24 46 36" fill="none" stroke="#C65D00" stroke-width="1.3"/>
<path d="M20 42 Q32 32 44 42" fill="none" stroke="#C65D00" stroke-width="1.3"/>
<path d="M29 18 C29 12 35 10 35 16 C35 10 39 12 39 18 C37 14 31 14 29 18Z" fill="#3D8B37"/>
<rect x="31" y="16" width="4" height="6" rx="1.5" fill="#5DA848"/>''',
    "onion": '''
<circle cx="32" cy="36" r="18" fill="#C9A0DC"/>
<circle cx="32" cy="36" r="14" fill="#B57EDC"/>
<circle cx="32" cy="36" r="10" fill="#A569BD"/>
<path d="M32 18 C34 14 36 12 32 10 C28 12 30 14 32 18Z" fill="#7DCEA0"/>
<path d="M20 36 C24 30 40 30 44 36" fill="none" stroke="#8E6BAF" stroke-width="1.2" opacity=".7"/>
<path d="M22 40 C26 34 38 34 42 40" fill="none" stroke="#8E6BAF" stroke-width="1.2" opacity=".7"/>''',
    "sugarcane": '''
<rect x="27" y="10" width="10" height="42" rx="5" fill="#6BBF59"/>
<rect x="27" y="16" width="10" height="2.5" fill="#4FA042"/>
<rect x="27" y="24" width="10" height="2.5" fill="#4FA042"/>
<rect x="27" y="32" width="10" height="2.5" fill="#4FA042"/>
<rect x="27" y="40" width="10" height="2.5" fill="#4FA042"/>
<ellipse cx="32" cy="12" rx="5" ry="3" fill="#8FD67A"/>''',
    "lavender": '''
<rect x="30" y="30" width="4" height="18" rx="2" fill="#7D6608"/>
<ellipse cx="24" cy="24" rx="5" ry="9" fill="#9B59B6"/>
<ellipse cx="32" cy="18" rx="5" ry="11" fill="#8E44AD"/>
<ellipse cx="40" cy="24" rx="5" ry="9" fill="#9B59B6"/>
<circle cx="24" cy="20" r="2" fill="#D7BDE2"/><circle cx="32" cy="14" r="2" fill="#D7BDE2"/>
<circle cx="40" cy="20" r="2" fill="#D7BDE2"/>''',
    "flour": '''
<path d="M16 20 H48 L44 50 H20 Z" fill="#F7F7F7" stroke="#D0D0D0" stroke-width="1.5"/>
<path d="M20 20 Q32 12 44 20" fill="#EFEFEF"/>
<ellipse cx="32" cy="26" rx="12" ry="5" fill="#FFFFFF"/>
<ellipse cx="32" cy="38" rx="10" ry="6" fill="#FFF8E7"/>
<circle cx="28" cy="36" r="1.5" fill="#FFF"/><circle cx="32" cy="39" r="1.5" fill="#FFF"/>
<circle cx="36" cy="36" r="1.5" fill="#FFF"/>''',
    "cornmeal": '''
<rect x="18" y="18" width="28" height="32" rx="3" fill="#F4D03F" stroke="#D4AC0D" stroke-width="1.5"/>
<rect x="22" y="22" width="20" height="12" rx="2" fill="#FFF"/>
<ellipse cx="32" cy="28" rx="6" ry="4" fill="#F1C40F"/>
<circle cx="26" cy="40" r="2" fill="#E67E22"/><circle cx="32" cy="42" r="2" fill="#E67E22"/>
<circle cx="38" cy="40" r="2" fill="#E67E22"/>''',
    "potato-chips": '''
<ellipse cx="22" cy="34" rx="12" ry="7" fill="#F1C40F" stroke="#E59866" stroke-width="1"/>
<ellipse cx="34" cy="28" rx="12" ry="7" fill="#F7DC6F" stroke="#E59866" stroke-width="1" transform="rotate(15 34 28)"/>
<ellipse cx="42" cy="38" rx="11" ry="6" fill="#F1C40F" stroke="#E59866" stroke-width="1" transform="rotate(-10 42 38)"/>
<ellipse cx="28" cy="30" rx="4" ry="2" fill="#FFF" opacity=".35"/>''',
    "ketchup": '''
<rect x="26" y="14" width="12" height="8" rx="2" fill="#922B21"/>
<path d="M24 22 H40 C40 22 38 50 32 50 C26 50 24 22 24 22Z" fill="#E74C3C"/>
<rect x="28" y="28" width="8" height="14" rx="1" fill="#FFF" opacity=".25"/>
<rect x="30" y="18" width="4" height="4" rx="1" fill="#C0392B"/>''',
    "chili-powder": '''
<path d="M18 18 H46 L42 50 H22 Z" fill="#FDF2E9" stroke="#E59866" stroke-width="1.5"/>
<ellipse cx="32" cy="24" rx="10" ry="4" fill="#FFF"/>
<circle cx="32" cy="36" r="10" fill="#E74C3C"/>
<circle cx="29" cy="33" r="2" fill="#C0392B" opacity=".5"/>
<circle cx="35" cy="38" r="1.5" fill="#C0392B" opacity=".5"/>''',
    "hot-sauce": '''
<rect x="28" y="12" width="8" height="6" rx="1.5" fill="#7B241C"/>
<rect x="25" y="18" width="14" height="30" rx="4" fill="#E74C3C" stroke="#922B21" stroke-width="1"/>
<rect x="28" y="24" width="8" height="16" rx="1" fill="#F9E79F"/>
<rect x="29" y="26" width="6" height="3" fill="#FFF" opacity=".4"/>''',
    "pickled-cucumber": '''
<rect x="20" y="16" width="24" height="34" rx="6" fill="#A9DFBF" stroke="#52BE80" stroke-width="1.5"/>
<rect x="20" y="16" width="24" height="10" rx="5" fill="#7DCEA0"/>
<ellipse cx="32" cy="34" rx="7" ry="12" fill="#58D68D"/>
<ellipse cx="32" cy="34" rx="4" ry="8" fill="#2ECC71" opacity=".5"/>''',
    "carrot-juice": '''
<rect x="28" y="12" width="8" height="8" rx="2" fill="#E67E22"/>
<rect x="26" y="20" width="12" height="30" rx="3" fill="#FDEBD0" stroke="#E59866" stroke-width="1"/>
<rect x="28" y="24" width="8" height="22" rx="2" fill="#F39C12"/>
<ellipse cx="32" cy="30" rx="3" ry="5" fill="#FFF" opacity=".25"/>''',
    "cereal": '''
<rect x="16" y="14" width="32" height="38" rx="3" fill="#5DADE2" stroke="#2E86C1" stroke-width="1.5"/>
<rect x="20" y="18" width="24" height="14" rx="2" fill="#FFF"/>
<ellipse cx="32" cy="25" rx="8" ry="4" fill="#F4D03F"/>
<circle cx="24" cy="38" r="2.5" fill="#E67E22"/><circle cx="32" cy="40" r="2.5" fill="#F4D03F"/>
<circle cx="40" cy="38" r="2.5" fill="#E67E22"/><circle cx="28" cy="44" r="2" fill="#F4D03F"/>''',
    "tofu": '''
<rect x="16" y="20" width="32" height="26" rx="3" fill="#FAFAFA" stroke="#D5D8DC" stroke-width="1.5"/>
<rect x="20" y="24" width="10" height="10" rx="1" fill="#F2F3F4"/>
<rect x="34" y="24" width="10" height="10" rx="1" fill="#F2F3F4"/>
<rect x="20" y="36" width="10" height="6" rx="1" fill="#F2F3F4"/>
<rect x="34" y="36" width="10" height="6" rx="1" fill="#F2F3F4"/>''',
    "dried-tofu": '''
<rect x="18" y="22" width="28" height="20" rx="2" fill="#D4A574" stroke="#A0714F" stroke-width="1.5"/>
<line x1="22" y1="27" x2="42" y2="27" stroke="#A0714F" stroke-width="1"/>
<line x1="22" y1="31" x2="42" y2="31" stroke="#A0714F" stroke-width="1"/>
<line x1="22" y1="35" x2="42" y2="35" stroke="#A0714F" stroke-width="1"/>
<line x1="22" y1="39" x2="42" y2="39" stroke="#A0714F" stroke-width="1"/>''',
    "peanut-butter": '''
<rect x="26" y="14" width="12" height="7" rx="2" fill="#A0714F"/>
<rect x="22" y="21" width="20" height="28" rx="4" fill="#F4D03F" stroke="#D4AC0D" stroke-width="1.5"/>
<ellipse cx="32" cy="34" rx="7" ry="9" fill="#D68910"/>
<ellipse cx="30" cy="32" rx="2" ry="3" fill="#B9770E" opacity=".5"/>''',
    "pumpkin-puree": '''
<rect x="22" y="18" width="20" height="30" rx="3" fill="#FFF" stroke="#D5D8DC" stroke-width="1.5"/>
<rect x="28" y="14" width="8" height="6" rx="2" fill="#D5D8DC"/>
<ellipse cx="32" cy="34" rx="7" ry="10" fill="#E67E22"/>
<ellipse cx="32" cy="32" rx="4" ry="6" fill="#F39C12" opacity=".5"/>''',
    "onion-soup-mix": '''
<rect x="18" y="20" width="28" height="30" rx="3" fill="#D7BDE2" stroke="#A569BD" stroke-width="1.5"/>
<rect x="22" y="24" width="20" height="10" rx="2" fill="#FFF"/>
<circle cx="26" cy="40" r="3" fill="#BB8FCE"/><circle cx="32" cy="42" r="3" fill="#A569BD"/>
<circle cx="38" cy="40" r="3" fill="#BB8FCE"/>''',
    "eggplant-slice": '''
<ellipse cx="32" cy="34" rx="20" ry="12" fill="#8E44AD" stroke="#6C3483" stroke-width="1.5"/>
<ellipse cx="32" cy="34" rx="14" ry="8" fill="#BB8FCE" opacity=".45"/>
<ellipse cx="32" cy="34" rx="5" ry="4" fill="#6C3483"/>
<ellipse cx="26" cy="32" rx="3" ry="2" fill="#FFF" opacity=".15"/>''',
    "eggplant-casserole": '''
<ellipse cx="32" cy="46" rx="20" ry="5" fill="#D5D8DC"/>
<path d="M14 38 Q32 28 50 38 L48 44 Q32 36 16 44Z" fill="#E59866" stroke="#CA6F1E" stroke-width="1"/>
<ellipse cx="32" cy="38" rx="14" ry="5" fill="#8E44AD"/>
<ellipse cx="32" cy="37" rx="10" ry="3" fill="#A569BD" opacity=".6"/>''',
    "watermelon-juice": '''
<rect x="28" y="12" width="8" height="7" rx="2" fill="#27AE60"/>
<rect x="26" y="19" width="12" height="32" rx="3" fill="#FDEBD0" stroke="#E59866" stroke-width="1"/>
<rect x="28" y="23" width="8" height="24" rx="2" fill="#E74C3C"/>
<circle cx="30" cy="30" r="1" fill="#2ECC71" opacity=".8"/>
<circle cx="34" cy="36" r="1" fill="#2ECC71" opacity=".8"/>''',
    "watermelon-smoothie": '''
<path d="M22 22 H42 L40 48 H24 Z" fill="#FADBD8" stroke="#E8A0A0" stroke-width="1.5"/>
<rect x="24" y="24" width="16" height="20" rx="2" fill="#E74C3C"/>
<rect x="28" y="14" width="12" height="10" rx="3" fill="#FFF" stroke="#D5D8DC" stroke-width="1"/>
<line x1="30" y1="16" x2="34" y2="22" stroke="#E74C3C" stroke-width="2"/>''',
    "essential-oil": '''
<rect x="29" y="12" width="6" height="5" rx="1" fill="#7D6608"/>
<rect x="27" y="17" width="10" height="6" rx="2" fill="#5D4037"/>
<rect x="26" y="23" width="12" height="24" rx="3" fill="#AF7AC5" stroke="#8E44AD" stroke-width="1"/>
<rect x="28" y="28" width="8" height="14" rx="1" fill="#D7BDE2"/>''',
    "perfume": '''
<rect x="29" y="10" width="6" height="5" rx="1" fill="#D4AC0D"/>
<ellipse cx="32" cy="18" rx="8" ry="3" fill="#F9E79F" stroke="#D4AC0D" stroke-width="1"/>
<path d="M24 18 H40 L38 50 H26 Z" fill="#FCF3CF" stroke="#D4AC0D" stroke-width="1.5"/>
<ellipse cx="32" cy="32" rx="6" ry="10" fill="#AF7AC5" opacity=".55"/>''',
    "blueberry": '''
<circle cx="24" cy="32" r="8" fill="#3498DB" stroke="#2471A3" stroke-width="1"/>
<circle cx="36" cy="30" r="8" fill="#5DADE2" stroke="#2471A3" stroke-width="1"/>
<circle cx="30" cy="40" r="8" fill="#2E86C1" stroke="#1B4F72" stroke-width="1"/>
<ellipse cx="22" cy="30" rx="2.5" ry="1.5" fill="#AED6F1" opacity=".7"/>
<ellipse cx="34" cy="28" rx="2.5" ry="1.5" fill="#AED6F1" opacity=".7"/>''',
    "strawberry-jam": '''
<rect x="22" y="22" width="20" height="26" rx="4" fill="#FFF" stroke="#D5D8DC" stroke-width="1.5"/>
<rect x="26" y="16" width="12" height="8" rx="2" fill="#E74C3C"/>
<rect x="26" y="28" width="12" height="14" rx="2" fill="#C0392B"/>
<ellipse cx="32" cy="34" rx="4" ry="5" fill="#E74C3C" opacity=".6"/>''',
    "blueberry-jam": '''
<rect x="22" y="22" width="20" height="26" rx="4" fill="#FFF" stroke="#D5D8DC" stroke-width="1.5"/>
<rect x="26" y="16" width="12" height="8" rx="2" fill="#3498DB"/>
<rect x="26" y="28" width="12" height="14" rx="2" fill="#2471A3"/>
<circle cx="29" cy="33" r="2" fill="#5DADE2" opacity=".6"/>
<circle cx="35" cy="36" r="2" fill="#5DADE2" opacity=".6"/>''',
    "apple-slice": '''
<path d="M18 34 C22 18 42 18 46 34 C42 44 22 44 18 34Z" fill="#E74C3C" stroke="#C0392B" stroke-width="1.5"/>
<path d="M32 20 V42" stroke="#FDFEFE" stroke-width="2"/>
<path d="M32 20 V42" stroke="#F5B7B1" stroke-width="8" opacity=".35"/>
<ellipse cx="32" cy="34" rx="5" ry="7" fill="#FADBD8"/>''',
    "apple-pie": '''
<ellipse cx="32" cy="44" rx="18" ry="5" fill="#D5D8DC"/>
<path d="M14 38 Q32 22 50 38 L48 44 Q32 32 16 44Z" fill="#E59866" stroke="#CA6F1E" stroke-width="1"/>
<path d="M18 38 Q32 28 46 38" fill="#FDFEFE"/>
<path d="M22 37 L28 30 L36 34 L42 28" stroke="#E74C3C" stroke-width="2" fill="none"/>''',
    "mango-juice": '''
<rect x="28" y="12" width="8" height="7" rx="2" fill="#F39C12"/>
<rect x="26" y="19" width="12" height="32" rx="3" fill="#FDEBD0" stroke="#E59866" stroke-width="1"/>
<rect x="28" y="23" width="8" height="24" rx="2" fill="#F4D03F"/>
<ellipse cx="32" cy="30" rx="3" ry="5" fill="#FFF" opacity=".25"/>''',
    "mango-smoothie": '''
<path d="M22 22 H42 L40 48 H24 Z" fill="#FDEBD0" stroke="#E59866" stroke-width="1.5"/>
<rect x="24" y="24" width="16" height="20" rx="2" fill="#F4D03F"/>
<rect x="28" y="14" width="12" height="10" rx="3" fill="#FFF" stroke="#D5D8DC" stroke-width="1"/>''',
    "grape-juice": '''
<rect x="28" y="12" width="8" height="7" rx="2" fill="#8E44AD"/>
<rect x="26" y="19" width="12" height="32" rx="3" fill="#FDEBD0" stroke="#E59866" stroke-width="1"/>
<rect x="28" y="23" width="8" height="24" rx="2" fill="#7D3C98"/>
<ellipse cx="32" cy="30" rx="3" ry="5" fill="#FFF" opacity=".2"/>''',
    "sugar": '''
<rect x="18" y="20" width="28" height="28" rx="3" fill="#FFF" stroke="#D5D8DC" stroke-width="1.5"/>
<rect x="22" y="24" width="20" height="10" rx="2" fill="#5DADE2"/>
<rect x="24" y="26" width="16" height="6" rx="1" fill="#AED6F1"/>
<circle cx="26" cy="40" r="2" fill="#FFF" stroke="#D5D8DC" stroke-width=".8"/>
<circle cx="32" cy="42" r="2" fill="#FFF" stroke="#D5D8DC" stroke-width=".8"/>
<circle cx="38" cy="40" r="2" fill="#FFF" stroke="#D5D8DC" stroke-width=".8"/>''',
    "oil": '''
<rect x="27" y="12" width="10" height="7" rx="2" fill="#D4AC0D"/>
<path d="M24 19 H40 C39 19 38 48 32 48 C26 48 25 19 25 19Z" fill="#F9E79F" stroke="#D4AC0D" stroke-width="1.5"/>
<ellipse cx="32" cy="34" rx="6" ry="10" fill="#F4D03F"/>
<ellipse cx="30" cy="30" rx="2" ry="4" fill="#FFF" opacity=".3"/>''',
}


def fetch_twemoji(code: str) -> str:
    url = TWEMOJI.format(code)
    last_error = None
    for _ in range(3):
        try:
            with urllib.request.urlopen(url, timeout=30) as resp:
                raw = resp.read().decode("utf-8")
            break
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    else:
        raise last_error

    inner = re.sub(r"<\?xml[^?]*\?>", "", raw)
    inner = re.sub(r"<!--.*?-->", "", inner, flags=re.S)
    m = re.search(r"<svg[^>]*>(.*)</svg>", inner, re.S)
    if not m:
        raise ValueError(f"invalid twemoji svg: {code}")
    body = m.group(1).strip()
    return (
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64" width="64" height="64">'
        f"{SVG_BG}"
        f'<g transform="translate(6,6) scale(1.444)">{body}</g>'
        "</svg>"
    )


def build_custom(slug: str) -> str:
    body = CUSTOM[slug]
    return (
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64" width="64" height="64">'
        f"{SVG_BG}{body}</svg>"
    )


def write_excel(rows: list[tuple]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "farm-icons"
    headers = ["序号", "名称", "icon名称", "icon路径"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 36
    for r in range(2, len(rows) + 2):
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    wb.save(EXCEL)


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    excel_rows = []

    for idx, (name, slug, emoji) in enumerate(ORDER, start=1):
        filename = f"{slug}.svg"
        icon_path = f"/images/farm/icons/{filename}"

        if emoji:
            try:
                svg = fetch_twemoji(emoji)
            except Exception as exc:  # noqa: BLE001
                if slug not in CUSTOM:
                    raise
                print(f"  warn: twemoji {emoji} failed ({exc}), use custom for {slug}")
                svg = build_custom(slug)
        else:
            svg = build_custom(slug)

        (OUTPUT / filename).write_text(svg.strip() + "\n", encoding="utf-8")
        excel_rows.append((idx, name, filename, icon_path))
        print(f"[{idx:02d}] {name} -> {filename}")

    write_excel(excel_rows)
    print(f"\nGenerated {len(ORDER)} icons -> {OUTPUT}")
    print(f"Excel -> {EXCEL}")


if __name__ == "__main__":
    main()
