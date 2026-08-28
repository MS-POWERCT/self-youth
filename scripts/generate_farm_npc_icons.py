#!/usr/bin/env python3
"""Generate recognizable transparent NPC SVG icons (64x64, chibi bust style)."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent / ".pylibs"))
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "public" / "images" / "farm" / "icons" / "npc"
EXCEL = ROOT / "public" / "images" / "farm" / "farm-npc-icons.xlsx"
ICON_BASE = "/images/farm/icons/npc"

HEADER = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64" width="64" height="64">'
FOOTER = "</svg>"

ORDER = [
    ("村长老外", "village-chief"),
    ("铁匠大锤", "blacksmith"),
    ("面包师小圆", "baker"),
    ("药师清心", "pharmacist"),
    ("裁缝锦绣", "tailor"),
    ("酒商老陈", "wine-merchant"),
    ("果农阿甘", "fruit-farmer"),
    ("渔夫阿海", "fisherman"),
    ("甜品师宣宣", "pastry-chef"),
    ("骑士守护", "knight"),
    ("炼金师摩根", "alchemist"),
    ("花匠小兰", "florist"),
    ("商人马克", "merchant"),
    ("御厨天王", "royal-chef"),
    ("世界树守护者", "world-tree-guardian"),
]

# 半身 Q 版头像：大脸 + 职业道具 + 描边，小尺寸也易辨认
ICONS = {
    "village-chief": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M16 44 Q32 34 48 44 L46 56 Q32 50 18 56Z" fill="#4A6741" stroke="#2E4A28" stroke-width="1"/>
  <circle cx="32" cy="28" r="13" fill="#F0C9A8" stroke="#C49A6C" stroke-width="1"/>
  <path d="M14 22 Q32 6 50 22 L48 28 Q32 14 16 28Z" fill="#8B5A2B" stroke="#5D3A1A" stroke-width="1"/>
  <rect x="16" y="24" width="32" height="5" rx="2" fill="#A0522D"/>
  <ellipse cx="32" cy="32" rx="9" ry="5" fill="#E8D5C4"/>
  <path d="M24 30 Q32 36 40 30" stroke="#8D6E63" stroke-width="1.2" fill="none"/>
  <circle cx="27" cy="26" r="1.6" fill="#333"/><circle cx="37" cy="26" r="1.6" fill="#333"/>
  <path d="M28 48 L28 58" stroke="#7D6608" stroke-width="2.5" stroke-linecap="round"/>
  <circle cx="28" cy="58" r="2" fill="#8B6914"/>''',

    "blacksmith": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#566573" stroke="#34495E" stroke-width="1"/>
  <rect x="20" y="40" width="24" height="14" rx="2" fill="#E67E22" stroke="#CA6F1E" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#D7A87E" stroke="#A0714F" stroke-width="1"/>
  <rect x="22" y="14" width="20" height="6" rx="2" fill="#7F8C8D"/>
  <circle cx="28" cy="25" r="1.5" fill="#222"/><circle cx="36" cy="25" r="1.5" fill="#222"/>
  <ellipse cx="32" cy="30" rx="2" ry="1" fill="#555" opacity=".5"/>
  <rect x="42" y="8" width="5" height="22" rx="2" fill="#7D6608" transform="rotate(25 44 19)"/>
  <rect x="44" y="4" width="12" height="7" rx="2" fill="#95A5A6" transform="rotate(25 50 7)"/>
  <rect x="10" y="46" width="18" height="6" rx="2" fill="#7F8C8D"/>
  <rect x="12" y="42" width="14" height="5" rx="1" fill="#566573"/>''',

    "baker": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#FDFEFE" stroke="#D5DBDB" stroke-width="1"/>
  <circle cx="32" cy="28" r="14" fill="#FAD7C5" stroke="#E0A98A" stroke-width="1"/>
  <path d="M18 16 Q32 4 46 16 L44 22 Q32 10 20 22Z" fill="#FFF" stroke="#D5DBDB" stroke-width="1"/>
  <circle cx="28" cy="26" r="1.5" fill="#333"/><circle cx="36" cy="26" r="1.5" fill="#333"/>
  <path d="M27 31 Q32 34 37 31" stroke="#E74C3C" stroke-width="1.2" fill="none"/>
  <ellipse cx="32" cy="30" rx="3" ry="2" fill="#F5B7B1" opacity=".5"/>
  <path d="M8 36 C12 30 18 28 24 32 C18 36 12 40 8 36Z" fill="#E59866" stroke="#CA6F1E" stroke-width="1"/>
  <path d="M10 34 C14 32 18 33 20 35" stroke="#CA6F1E" stroke-width="1" fill="none"/>''',

    "pharmacist": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#FDFEFE" stroke="#AED6F1" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#F5CBA7" stroke="#D4A574" stroke-width="1"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <path d="M29 30 Q32 32 35 30" stroke="#C0392B" stroke-width="1" fill="none"/>
  <rect x="40" y="18" width="12" height="28" rx="4" fill="#58D68D" stroke="#27AE60" stroke-width="1"/>
  <rect x="43" y="22" width="6" height="6" rx="1" fill="#FFF"/>
  <line x1="46" y1="23" x2="46" y2="27" stroke="#E74C3C" stroke-width="1.5"/>
  <line x1="44" y1="25" x2="48" y2="25" stroke="#E74C3C" stroke-width="1.5"/>
  <ellipse cx="46" cy="36" rx="3" ry="6" fill="#ABEBC6" opacity=".7"/>''',

    "tailor": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#AF7AC5" stroke="#7D3C98" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#FADBD8" stroke="#E8A0A0" stroke-width="1"/>
  <path d="M22 18 Q32 10 42 18" fill="#5D4037" stroke="#3E2723" stroke-width="1"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <path d="M10 28 L18 36 L14 40 L6 32Z" fill="#95A5A6" stroke="#566573" stroke-width="1"/>
  <path d="M54 28 L46 36 L50 40 L58 32Z" fill="#95A5A6" stroke="#566573" stroke-width="1"/>
  <circle cx="32" cy="38" r="3" fill="#E74C3C"/>
  <ellipse cx="48" cy="46" rx="5" ry="4" fill="#F5B7B1" stroke="#E8A0A0" stroke-width="1"/>
  <circle cx="48" cy="46" r="2" fill="#FFF" opacity=".5"/>''',

    "wine-merchant": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#922B21" stroke="#641E16" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#E8C4A0" stroke="#C49A6C" stroke-width="1"/>
  <path d="M22 18 Q32 12 42 18" fill="#BDC3C7" stroke="#7F8C8D" stroke-width="1"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <path d="M27 30 Q32 32 37 30" stroke="#8D6E63" stroke-width="1" fill="none"/>
  <path d="M44 16 H52 L50 44 H46 Z" fill="#6C3483" stroke="#4A235A" stroke-width="1"/>
  <rect x="47" y="12" width="6" height="6" rx="2" fill="#922B21"/>
  <ellipse cx="48" cy="30" rx="2" ry="5" fill="#BB8FCE" opacity=".5"/>
  <ellipse cx="8" cy="40" rx="4" ry="3" fill="#F5EEF8" stroke="#D7BDE2" stroke-width="1"/>''',

    "fruit-farmer": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#27AE60" stroke="#1E8449" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#F0C9A8" stroke="#C49A6C" stroke-width="1"/>
  <ellipse cx="32" cy="14" rx="16" ry="5" fill="#E59866" stroke="#CA6F1E" stroke-width="1"/>
  <rect x="16" y="14" width="32" height="3" rx="1" fill="#D68910"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <path d="M29 30 Q32 32 35 30" stroke="#C0392B" stroke-width="1" fill="none"/>
  <circle cx="46" cy="38" r="7" fill="#E74C3C" stroke="#C0392B" stroke-width="1"/>
  <ellipse cx="44" cy="35" rx="2" ry="3" fill="#27AE60"/>
  <circle cx="8" cy="42" r="5" fill="#F4D03F" stroke="#D4AC0D" stroke-width="1"/>''',

    "fisherman": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#3498DB" stroke="#2471A3" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#F0C9A8" stroke="#C49A6C" stroke-width="1"/>
  <path d="M16 18 Q32 8 48 18 L46 24 Q32 14 18 24Z" fill="#F4D03F" stroke="#D4AC0D" stroke-width="1"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <path d="M6 14 Q20 30 14 52" stroke="#7D6608" stroke-width="2" fill="none" stroke-linecap="round"/>
  <path d="M12 48 C18 44 28 42 38 46 C32 50 22 52 12 48Z" fill="#F39C12" stroke="#E67E22" stroke-width="1"/>
  <circle cx="16" cy="47" r="1.5" fill="#FFF" opacity=".6"/>
  <circle cx="34" cy="46" r="1.5" fill="#333"/>''',

    "pastry-chef": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#F5B7B1" stroke="#E8A0A0" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#FADBD8" stroke="#E8A0A0" stroke-width="1"/>
  <path d="M20 14 Q32 6 44 14 L42 20 Q32 12 22 20Z" fill="#FFF" stroke="#F5B7B1" stroke-width="1"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <path d="M29 30 Q32 32 35 30" stroke="#E74C3C" stroke-width="1" fill="none"/>
  <rect x="42" y="34" width="14" height="12" rx="3" fill="#F1948A" stroke="#E74C3C" stroke-width="1"/>
  <rect x="44" y="30" width="10" height="6" rx="2" fill="#FFF" stroke="#F5B7B1" stroke-width="1"/>
  <circle cx="49" cy="29" r="2" fill="#E74C3C"/>
  <circle cx="46" cy="40" r="1.5" fill="#FFF" opacity=".7"/>''',

    "knight": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M22 18 H42 L40 34 H24 Z" fill="#AAB7B8" stroke="#566573" stroke-width="1"/>
  <rect x="26" y="34" width="12" height="4" rx="1" fill="#566573"/>
  <rect x="20" y="38" width="24" height="18" rx="3" fill="#5D6D7E" stroke="#34495E" stroke-width="1"/>
  <rect x="24" y="42" width="16" height="10" rx="1" fill="#85929E"/>
  <rect x="28" y="44" width="8" height="6" fill="#F4D03F" stroke="#D4AC0D" stroke-width=".8"/>
  <path d="M30 22 H34 V28 H30Z" fill="#212F3D" opacity=".6"/>
  <rect x="8" y="28" width="10" height="14" rx="2" fill="#7F8C8D" stroke="#566573" stroke-width="1"/>
  <rect x="10" y="32" width="6" height="6" fill="#3498DB" stroke="#2471A3" stroke-width=".8"/>''',

    "alchemist": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#6C3483" stroke="#4A235A" stroke-width="1"/>
  <circle cx="32" cy="26" r="11" fill="#E8C4A0" stroke="#C49A6C" stroke-width="1"/>
  <path d="M24 10 L32 4 L40 10 L38 18 H26 Z" fill="#4A235A" stroke="#2E1054" stroke-width="1"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <path d="M44 20 H52 L50 46 H46 Z" fill="#D7BDE2" stroke="#8E44AD" stroke-width="1"/>
  <rect x="47" y="16" width="6" height="6" rx="2" fill="#8E44AD"/>
  <circle cx="48" cy="30" r="3" fill="#58D68D" opacity=".8"/>
  <circle cx="46" cy="24" r="1.5" fill="#85C1E9" opacity=".9"/>
  <circle cx="50" cy="26" r="1" fill="#F4D03F" opacity=".9"/>''',

    "florist": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#58D68D" stroke="#27AE60" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#FADBD8" stroke="#E8A0A0" stroke-width="1"/>
  <path d="M22 16 Q32 8 42 16" fill="#5D4037" stroke="#3E2723" stroke-width="1"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <circle cx="24" cy="14" r="4" fill="#F1948A" stroke="#E74C3C" stroke-width=".8"/>
  <circle cx="32" cy="10" r="4" fill="#F5B7B1" stroke="#E8A0A0" stroke-width=".8"/>
  <circle cx="40" cy="14" r="4" fill="#F9E79F" stroke="#F4D03F" stroke-width=".8"/>
  <ellipse cx="48" cy="42" rx="6" ry="8" fill="#82E0AA" stroke="#27AE60" stroke-width="1"/>
  <circle cx="46" cy="38" r="3" fill="#F1948A"/>
  <circle cx="50" cy="42" r="3" fill="#F5B7B1"/>''',

    "merchant": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#1E8449" stroke="#145A32" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#F0C9A8" stroke="#C49A6C" stroke-width="1"/>
  <rect x="22" y="12" width="20" height="8" rx="2" fill="#212F3D" stroke="#17202A" stroke-width="1"/>
  <rect x="24" y="10" width="16" height="3" rx="1" fill="#212F3D"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <path d="M29 30 Q32 32 35 30" stroke="#8D6E63" stroke-width="1" fill="none"/>
  <ellipse cx="10" cy="42" rx="7" ry="8" fill="#D68910" stroke="#B9770E" stroke-width="1"/>
  <path d="M8 38 H12" stroke="#F4D03F" stroke-width="1"/>
  <circle cx="10" cy="42" r="3" fill="#F4D03F" opacity=".6"/>''',

    "royal-chef": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#922B21" stroke="#641E16" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#F0C9A8" stroke="#C49A6C" stroke-width="1"/>
  <path d="M20 14 Q32 6 44 14 L42 20 Q32 12 22 20Z" fill="#FFF" stroke="#D5DBDB" stroke-width="1"/>
  <path d="M24 10 H40 L38 14 H26 Z" fill="#F4D03F" stroke="#D4AC0D" stroke-width="1"/>
  <circle cx="28" cy="12" r="2" fill="#E74C3C"/><circle cx="32" cy="10" r="2" fill="#E74C3C"/>
  <circle cx="36" cy="12" r="2" fill="#E74C3C"/>
  <circle cx="28" cy="25" r="1.5" fill="#333"/><circle cx="36" cy="25" r="1.5" fill="#333"/>
  <ellipse cx="50" cy="36" rx="3" ry="6" fill="#F4D03F" stroke="#D4AC0D" stroke-width="1"/>
  <circle cx="50" cy="30" rx="4" ry="3" fill="#85929E" stroke="#566573" stroke-width="1"/>''',

    "world-tree-guardian": '''
  <ellipse cx="32" cy="52" rx="18" ry="8" fill="#000" opacity=".08"/>
  <path d="M18 44 Q32 34 46 44 L44 56 Q32 50 20 56Z" fill="#1E5631" stroke="#145A32" stroke-width="1"/>
  <circle cx="32" cy="26" r="12" fill="#D5F5E3" stroke="#82E0AA" stroke-width="1"/>
  <circle cx="28" cy="24" r="1.5" fill="#1E8449"/><circle cx="36" cy="24" r="1.5" fill="#1E8449"/>
  <path d="M29 28 Q32 30 35 28" stroke="#27AE60" stroke-width="1" fill="none"/>
  <rect x="44" y="20" width="4" height="28" rx="2" fill="#7D6608"/>
  <circle cx="46" cy="16" r="10" fill="#27AE60" stroke="#1E8449" stroke-width="1"/>
  <circle cx="42" cy="14" r="5" fill="#58D68D"/>
  <circle cx="50" cy="14" r="5" fill="#58D68D"/>
  <circle cx="46" cy="10" r="4" fill="#82E0AA"/>
  <circle cx="46" cy="16" r="3" fill="#F4D03F" opacity=".85"/>
  <ellipse cx="32" cy="8" rx="14" ry="4" fill="#F4D03F" opacity=".2"/>''',
}


def write_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "farm-npc-icons"
    headers = ["序号", "NPC名称", "icon名称", "icon路径"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="4472C4")
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 42
    wb.save(EXCEL)


def main():
    OUTPUT.mkdir(parents=True, exist_ok=True)
    rows = []
    for idx, (name, slug) in enumerate(ORDER, start=1):
        filename = f"{slug}.svg"
        svg = HEADER + ICONS[slug] + FOOTER
        (OUTPUT / filename).write_text(svg.strip() + "\n", encoding="utf-8")
        rows.append((idx, name, filename, f"{ICON_BASE}/{filename}"))
        print(f"[{idx:02d}] {name} -> {filename}")
    write_excel(rows)
    print(f"\nRegenerated {len(ORDER)} icons -> {OUTPUT}")


if __name__ == "__main__":
    main()
